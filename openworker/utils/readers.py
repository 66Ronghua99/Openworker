import os
from pathlib import Path
from typing import Optional
import pypdf
import docx
import openpyxl

def read_file_content(file_path: str) -> str:
    """
    Reads the content of various file types and returns a string.
    Supported types: .pdf, .docx, .xlsx, various text formats.
    """
    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found at {file_path}"
    
    suffix = path.suffix.lower()
    
    try:
        if suffix == ".pdf":
            return _read_pdf(path)
        elif suffix == ".docx":
            return _read_docx(path)
        elif suffix == ".xlsx":
            return _read_excel(path)
        else:
            # specialized handling or fallback to text
            return _read_text(path)
    except Exception as e:
        return f"Error reading file {file_path}: {str(e)}"

def _read_pdf(path: Path) -> str:
    text = []
    with open(path, "rb") as f:
        reader = pypdf.PdfReader(f)
        for page in reader.pages:
            text.append(page.extract_text() or "")
    return "\n".join(text)

def _read_docx(path: Path) -> str:
    doc = docx.Document(path)
    return "\n".join([para.text for para in doc.paragraphs])

def _read_excel(path: Path) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    text = []
    for sheet in wb.sheetnames:
        text.append(f"--- Sheet: {sheet} ---")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            # Convert row values to string and join
            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
            text.append(row_text)
    return "\n".join(text)

def _read_text(path: Path) -> str:
    # Try common encodings
    encodings = ["utf-8", "latin-1", "ascii"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    return "Error: Could not decode file with supported encodings."
